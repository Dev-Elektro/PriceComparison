from loguru import logger as log
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string as column_index
from searchengine import SearchPool, QueryItem, writeToFile
from searchengine.presetsite import dnsshop, regard, citilink
from multiprocessing import cpu_count

from web_main import website


def print_log(name: str, total_count: int, current_pos: int):
    log.info(f"{name} {total_count}\\{current_pos}")


def main():
    file_path = 'D:/MyProject/reestr.xlsx'
    file = load_workbook(file_path)  # Загружаем в память книгу xlsx
    sheet = file['Реестр']  # Открываем лист книги
    # Выборка всех ячеек с запросами из столбца 'D' начиная с 6 строки
    cells = list(filter(lambda x: x.data_type == 's', sheet['D'][6:]))
    search_query_dns = []
    search_query_regard = []
    search_query_citilink = []
    for cell in cells:  # Бежим по столбцу необходимых позиций
        # Получаем ячейку примечание с текущей строки
        cell_footnote = file['Запрос КП1'].cell(row=cell.row, column=column_index('H'))
        # Берем значение со столбца примечание, если не заполнено берем значение со столбца необходимых позиций
        query = cell_footnote.value if cell_footnote.data_type == 's' else cell.value
        # Проверяем требуется запрос для dns-shop
        if check_for_wait_query(sheet, cell.row, 'K'):
            search_query_dns.append(QueryItem(rowNum=cell.row, value=query))
        # Проверяем требуется запрос для regard
        if check_for_wait_query(sheet, cell.row, 'L'):
            search_query_regard.append(QueryItem(rowNum=cell.row, value=query))
        # Проверяем требуется запрос для regard
        if check_for_wait_query(sheet, cell.row, 'M'):
            search_query_citilink.append(QueryItem(rowNum=cell.row, value=query))
    search_pool = SearchPool(processes=cpu_count(), headless=False)
    search_pool.setCallback(print_log)
    search_pool.addTask('DNS-shop', dnsshop, search_query_dns, 'Реестр', 'K')
    search_pool.addTask('Regard', regard, search_query_regard, 'Реестр', 'L')
    search_pool.addTask('Citilink', citilink, search_query_citilink, 'Реестр', 'M')
    search_pool.setCallback(print_log)
    for searchPoolItem in search_pool:
        log.info(f"Search finish: {searchPoolItem.name} - write to column: {searchPoolItem.columnName}")
        for resultItem in searchPoolItem.result:  # Получаем результат выполнения поиска по сайту
            result_txt = "\n".join(map(lambda x: f"{x.name} - Цена: {x.price}",
                                       resultItem.listProduct))  # Формируем строку результата поиска.
            wcell = sheet.cell(row=resultItem.rowNum, column=column_index(
                searchPoolItem.columnName))  # Получаем ячейку в книге для записи результата.
            wcell.font = Font(name='Times New Roman', size=11)
            sheet.row_dimensions[resultItem.rowNum].height = 70
            wcell.value = result_txt if result_txt else 'Нет'  # Записываем результат в ячейку
    file.save(file_path)


@writeToFile
def write_file(site_name, sheet_name, column_name, result_item):
    file_path = 'D:/MyProject/reestr.xlsx'
    file = load_workbook(file_path)  # Загружаем в память книгу xlsx
    log.info(f"Запись в файла для сайта: {site_name}, книга: {sheet_name}, колонка: {column_name}")
    sheet = file[sheet_name]  # Открываем лист книги
    # Формируем строку результата поиска.
    result_txt = "\n".join(map(lambda x: f"{x.name} - Цена: {x.price}", result_item.listProduct))
    # Получаем ячейку в книге для записи результата.
    wcell = sheet.cell(row=result_item.rowNum, column=column_index(column_name))
    wcell.font = Font(name='Times New Roman', size=11)
    sheet.row_dimensions[result_item.rowNum].height = 70
    wcell.value = result_txt if result_txt else 'Нет'  # Записываем результат в ячейку
    file.save(file_path)
    log.info(f"Файл сохранен.")


def check_for_wait_query(sheet, row: int, column: str):
    return sheet.cell(row=row, column=column_index(column)).data_type == 'n'


def main2():
    file_path = 'D:/MyProject/reestr.xlsx'
    file = load_workbook(file_path)  # Загружаем в память книгу xlsx
    sheet = file['Реестр']  # Открываем лист книги
    # Выборка всех ячеек с запросами из столбца 'D' начиная с 6 строки
    cells = list(filter(lambda x: x.data_type == 's', sheet['E'][6:]))
    search_query_dns = []
    search_query_regard = []
    for cell in cells:  # Бежим по столбцу необходимых позиций
        # Получаем ячейку примечание с текущей строки
        cell_footnote = file['Запрос КП1'].cell(row=cell.row, column=column_index('H'))
        # Берем значение со столбца примечание, если не заполнено берем значение со столбца необходимых позиций
        query = cell_footnote.value if cell_footnote.data_type == 's' else cell.value
        # Проверяем требуется запрос для dns-shop
        if check_for_wait_query(sheet, cell.row, 'K'):
            search_query_dns.append(QueryItem(rowNum=cell.row, value=query))
        # Проверяем требуется запрос для regard
        if check_for_wait_query(sheet, cell.row, 'L'):
            search_query_regard.append(QueryItem(rowNum=cell.row, value=query))
    search_pool = SearchPool(processes=cpu_count(), headless=False)
    search_pool.setCallback(print_log)
    search_pool.addTask('DNS-shop', dnsshop, search_query_dns, 'Реестр', 'K')
    search_pool.addTask('Regard', regard, search_query_regard, 'Реестр', 'L')
    file.close()
    search_pool.wait()


if __name__ == '__main2__':
    import sys
    if len(sys.argv) == 1:
        website("one")
    elif sys.argv[1] == 'registry':
        website("reestr")
