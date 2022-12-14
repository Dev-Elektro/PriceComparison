from datetime import datetime
from distutils.log import error
from enum import auto
import os
from time import sleep
from tkinter.messagebox import YESNOCANCEL
from searchengine.presetsite import citilink, regard, dnsshop
from searchengine.webdriver import Driver
from searchengine import runSearchBySite
from cgitb import text
from tkinter import filedialog
from tkinter import *
from tkinter.filedialog import askopenfilename
from numpy import append
import subprocess
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import eel
from random import randint
import shutil

driver = Driver(headless = False)

def write_file_index(index):
      f = open('index.txt','w')
      f.write(index)
      f.close()

def write_file_index_nach(index):
      f = open('index_0.txt','w')
      f.write(index)
      f.close()

def read_file_index():
      f = open('index.txt','r')
      rindex=f.read()
      f.close()
      return rindex

def write_file_index_kp(index):
      f = open('index_kp.txt','w') 
      for  line in index:
            f.write(line + '\n')
      f.close()

def read_file_index_kp():
      f = open('index_kp.txt','r')
      rindex = [line.rstrip('\n') for line in f]
      f.close()
      return rindex



#(путь к файлу, имя листа, имя  резервного листа, буква основной колонки, буква резервной колонки, имя столбца основного, имя резервного стобца, имя функции поиска цены, буква столбца записи, имя столбца записи)
@eel.expose
def read2_xlsx(xlsx, sheet_name,sheet_name_reserv, column_number,reserv_column, name_of_cell, reserv_name_of_cell, function_search, name_sheet_write, name_write, name_col_write):
      """Основная функция с алгоритмом обработки файла реестра и запуска поиска по поставщикам"""
      driver.start()
      eel.js_wait()
      name_of_magaz=function_search.__name__
      name_of_magaz=name_of_magaz.replace("searchSite.","")
      eel.my_javascript_function(f"• Запустили поиск по магазину {name_of_magaz} по листу {sheet_name}")
      check=False
      if not os.path.exists("index.txt"):
        open("index.txt", 'w').close()
      #xlsx='E:\\D disk\WorK\\bat source and compile\\Python\\priceZADA4a\\Итоговый реестр ШАБЛОН_2.xlsx'
      #sheet_name="Отчет"
      #column_number="h"
      #name_of_cell="Примечание"
      list_result=[]
      wb = load_workbook(xlsx)
      ws = wb[sheet_name_reserv]
      column=ws[reserv_column]
      
      #проверка наименования столбца
      for cell in column:
            if cell.value==reserv_name_of_cell:
                  print(f"Имя стартового столбца={cell.value}")
                  check=True
                  coordinatenachcalo=cell.row
                  print(f"{coordinatenachcalo=}")
                  write_file_index_nach(str(coordinatenachcalo))
                  coordinate_index_from_file=read_file_index()
                  if coordinate_index_from_file:
                        coordinate_now=coordinate_index_from_file
                  else:
                        coordinate_now=int(coordinatenachcalo)+1
                        write_file_index(str(coordinate_now))
      dlina=0
      dlina_str=0
      for i in column:
            if i.value:
                  dlina+=1
      print(f"Длина {dlina}")
      dlina=int(dlina)+int(coordinatenachcalo)
      print(f"Будет поиск до {dlina} строки")
      #считывание данных
      ws = wb[sheet_name]
      column=ws[column_number]
      for index,cell in  enumerate(column):
            #print(f"индекс таков {index}")
            if index==dlina:
                  break
            if check==True:
                        #print(cell.value)
                        index_of_table=read_file_index()
                        if index_of_table:
                              coordinate_now=int(index_of_table)
                        
                        if cell.row >= coordinate_now:
                              letter0=cell.column_letter
                              print(f"Смотрим ячейку {letter0}{coordinate_now}")
                              #list_result.extend([[cell.value,coordinate0]])
                              #print(list_result)
                              #result_search=search(cell.value)
                              yacheika_main=str(column_number)+str(cell.row)
                              ws = wb[sheet_name]
                              plan_main=ws[yacheika_main]
                              ws_check = wb[name_sheet_write]
                              yacheika_check=str(name_write)+str(cell.row)
                              plan_check=ws_check[yacheika_check] #проверка пустой ячейки для записи
                              print(f"{yacheika_check=}")
                                                                                         
                              if plan_main.value and not plan_check.value:
                                    print(f"{plan_main.value=}")
                                    print (f"{yacheika_main=}")
                                    buf = runSearchBySite(plan_main.value,function_search(driver))
                                    citilink_result=""
                                    list_res=[]
                                    for index, i in enumerate(buf):
                                          if index==4:
                                                break
                                          
                                          print(f"Вывод функции поиска {i.name} - Цена: {i.price}\n")
                                          list_res.extend([[i.name,i.price]])
                                    #print(f"Индекс для записи = {index}")
                                    #print(list_res)
                                    #print(len(list_res))
                                    dlina_str=len(list_res)

                                    if dlina_str>1:
                                          for s in list_res:
                                                citilink_result+=s[0]+" Цена: "+s[1]+"\n"
                                    else:
                                          if list_res:
                                                citilink_result=list_res[0][1]                                    
                                    print(f"{citilink_result=}")                                   

                                                               
                                    if citilink_result:
                                          result_search=str(citilink_result)
                                          write2_xlsx(file_put,name_sheet_write,name_write,name_col_write,result_search,coordinate_now,dlina_str)
                                          coordinate_now=int(coordinate_now)+1
                                          write_file_index(str(coordinate_now))
                                          #print(f"zapis {result_search}")
                                    else:
                                          result_search="Нет"
                                          write2_xlsx(file_put,name_sheet_write,name_write,name_col_write,result_search,coordinate_now,dlina_str)
                                          coordinate_now=int(coordinate_now)+1
                                          write_file_index(str(coordinate_now))
                                          #print(f"zapis {result_search}")

                              else:
                                    if not plan_check.value:
                                          print("*Ищем по резервному названию*")
                                          yacheika=str(reserv_column)+str(cell.row)
                                          ws = wb[sheet_name_reserv]
                                          planB=ws[yacheika]
                                    
                                          if planB.value:
                                                print(f"{planB.value=}")
                                                print(f"{yacheika=}")
                                                buf = runSearchBySite(planB.value,function_search(driver))
                                                citilink_result=""
                                                list_res=[]                                          

                                                for index, i in enumerate(buf):
                                                      if index==4:
                                                            break
                                                
                                                      print(f"Вывод функции поиска {i.name} - Цена: {i.price}\n")
                                                      list_res.extend([[i.name,i.price]])
                                                
                                                #print(list_res)
                                                #print(len(list_res))
                                                #print(f"Индекс для записи = {index}")
                                                dlina_str=len(list_res)

                                                if dlina_str>1:
                                                      for s in list_res:
                                                            citilink_result+=s[0]+" Цена: "+s[1]+"\n"
                                                else:
                                                      if list_res:
                                                            citilink_result=list_res[0][1]
                                                print(f"{citilink_result=}")     
                                          
                                                #eel.my_javascript_function(citilink_result)
                                                if citilink_result:
                                                      result_search=str(citilink_result)
                                                      write2_xlsx(file_put,name_sheet_write,name_write,name_col_write,result_search,coordinate_now,dlina_str)
                                                      coordinate_now=int(coordinate_now)+1
                                                      write_file_index(str(coordinate_now))
                                                      #print(f"zapis {result_search}")
                                                else:
                                                      result_search="Нет"
                                                      write2_xlsx(file_put,name_sheet_write,name_write,name_col_write,result_search,coordinate_now,dlina_str)
                                                      coordinate_now=int(coordinate_now)+1
                                                      write_file_index(str(coordinate_now))
                                                      #print(f"zapis {result_search}")
                                    

      write_file_index("") #очищаем файл с индексом
      write_file_index_nach("") #очищаем файл с начальным индексом
      print(f" Выполнен поиск по магазину {name_of_magaz}")
      eel.my_javascript_function("<font color=\"green\">"+f"✓ Завершен поиск по магазину {name_of_magaz} по листу {sheet_name}"+"</font>")
      eel.js_gotovo()
      driver.stop()
      


#(путь к файлу, имя листа, буква основной колонки, имя столбца основного, данные для записи, номер ячейки, длина)
def write2_xlsx(xlsx, sheet_name, column_number, name_of_cell,list_zapis,nomer2,dlina_str):
      """Функция записи в файл реестра (поиск по реестру)"""
      check=False
      #xlsx='E:\\D disk\WorK\\bat source and compile\\Python\\priceZADA4a\\Итоговый реестр ШАБЛОН_2.xlsx'
      #sheet_name="Отчет"
      #column_number="h"
      #name_of_cell="Примечание"
      #list_zapis=['ergegdyyhnjhju','cccccccc','ddddddddd','aaaaaaaaa','dddddddd','aaaaaaaaaaa','eddvegeggherhe']

      wb = load_workbook(xlsx)
      ws = wb[sheet_name]
      column=ws[column_number]
      ws["I6"]="Ситилинк"
      ws["J6"]="Регард"
      ws["K6"]="ДНС"
      for col in ['I', 'J', 'K']:
            if ws.column_dimensions[col].hidden == False:
                  ws.column_dimensions[col].hidden= True

      for cell in column:
            if cell.value==name_of_cell:
                  coordinate=cell.row
                  letter=cell.column_letter
                  print(f"Проверка столбца записи = {letter} {coordinate}")
                  check=True
      #nomer=int(coordinate)+1

      if check==True:
            #nomer2=read_file_index()
            yacheyka=letter+str(nomer2)
            print(f"Пишем в ячейку {yacheyka}")
            if list_zapis=="Нет":
                 ws[yacheyka]=list_zapis
                 megre_cell =ws[yacheyka]
                 megre_cell.fill = PatternFill('solid', fgColor="ffd694")
                 megre_cell.font = Font(name='Times New Roman', size=11)
                 rd = ws.row_dimensions[int(nomer2)] # get dimension for row
                 rd.height=70
            else:
                  if dlina_str>1:
                        print(f"Длина запись {dlina_str}")
                        megre_cell =ws[yacheyka]
                        ws[yacheyka]=list_zapis
                        megre_cell.fill = PatternFill('solid', fgColor="ff9494")
                        megre_cell.font = Font(name='Times New Roman', size=11)
                        rd = ws.row_dimensions[int(nomer2)] # get dimension for row
                        rd.height=70
                  else:
                        print(f" Длина запись {dlina_str}")
                        megre_cell =ws[yacheyka]
                        ws[yacheyka]=int(list_zapis)
                        megre_cell.fill = PatternFill('solid', fgColor="ffffff")
                        megre_cell.font = Font(name='Times New Roman', size=11)
                        rd = ws.row_dimensions[int(nomer2)] # get dimension for row
                        rd.height=70

      wb.save(xlsx)
      print(f"Выполнена запись в файл")
      




@eel.expose
def fileopen():
      """Функция выбора файла реестра для обработки"""
      # Exposing the random_python function to javascript   
      root = Tk()
      root.attributes('-toolwindow', True)
      root.eval('tk::PlaceWindow . center')
      root.withdraw()
      root.attributes("-topmost", True)

      global file_put     
      file_put=askopenfilename(defaultextension=".xslx",filetypes = [("Книга Excel","*.xlsx")])
      root.destroy()
      print(file_put)
      if(file_put):                   
            thisFile = file_put
            base = os.path.splitext(thisFile)[0]
            suffix=os.path.splitext(thisFile)[1]
            dst_file=base+"_ОБРАБОТАН"+suffix
            print(dst_file)
            shutil.copyfile(file_put,dst_file)
            #делаем копированный и переименованнный файл основным
            file_put=dst_file
            eel.my_javascript_function(f"✓ Сделана копия и обработается файл реестра: {file_put}") 
            return thisFile
      
#запись в файл при единичном поиске
@eel.expose
def write_toExcel():
      """Запись в файл после поиска по единичной позиции"""
      root = Tk()
      root.attributes('-toolwindow', True)
      root.eval('tk::PlaceWindow . center')
      root.withdraw()
      root.attributes("-topmost", True)
      time_write=datetime.now()
      time_write=time_write.strftime("%d.%m.%Y_%H_%M")
      root.ass=filedialog.asksaveasfilename(initialdir = "/", initialfile=f"Выгрузка_{time_write}.xlsx",title = "Выбирите куда сохранить",filetypes = (("Книга Excel","*.xlsx"),("all files","*.*")))
      file_put_write=root.ass
      print(file_put_write)
      wb_write = openpyxl.Workbook() 
      ws_write = wb_write.active
      ws_write.title="Разовый поиск"


               #A         B           C       D         E      F      G
      shapka=['№ п/п', 'Citilink', 'Цена', 'Regard', 'Цена' ,'DNS', 'Цена']
      ws_write.append(shapka)
      bold_col=["A1","B1","C1","D1","E1","F1","G1"]
      for col in bold_col:
            ws_write[col].font=Font(size = "12", bold=True)
            
      
      dlina_citilink=len(citilink_toex)
      dlina_regard=len(regard_toex)
      dlina_dns=len(dns_toex)
      print(f"{dlina_citilink=} {dlina_regard=} {dlina_dns=}")
      citilink_mass=""
      regard_mass=""
      dns_mass=""
      if citilink_toex:
            if dlina_citilink==1:
                  ws_write["B2"]=citilink_toex[0][0]
                  ws_write["C2"]=int(citilink_toex[0][1])
                  ws_write.column_dimensions["B"].width = 70
            else:
                  for cit in citilink_toex:
                        citilink_mass+=cit[0]+" -Цена:"+cit[1]+"\n"
                  ws_write["B2"]=citilink_mass
                  ws_write["B2"].alignment = Alignment(wrapText=True)
                  #ws_write["B2"].alignment = Alignment(horizontal="center", vertical="center")
                  ws_write.column_dimensions["B"].width = 70
                  ws_write.row_dimensions[2].height = 60

      if regard_toex:
            if dlina_regard==1:
                  ws_write["D2"]=regard_toex[0][0]
                  ws_write["E2"]=int(regard_toex[0][1])
                  ws_write.column_dimensions["D"].width = 70
            else:
                  for reg in regard_toex:
                        regard_mass+=reg[0]+" -Цена:"+reg[1]+"\n"
                  ws_write["D2"]=regard_mass
                  ws_write["D2"].alignment = Alignment(wrapText=True)
                 #ws_write["D2"].alignment = Alignment(horizontal="center", vertical="center")
                  ws_write.column_dimensions["D"].width = 70
                  ws_write.row_dimensions[2].height = 60
                        

      if dns_toex:
            if dlina_dns==1:
                  ws_write["F2"]=dns_toex[0][0]
                  ws_write["G2"]=int(dns_toex[0][1])
                  ws_write.column_dimensions["F"].width = 70
            else:
                  for dns in dns_toex:
                        dns_mass+=dns[0]+" -Цена:"+dns[1]+"\n"
                  ws_write["F2"]=dns_mass
                  ws_write["F2"].alignment = Alignment(wrapText=True)
                  #ws_write["F2"].alignment = Alignment(horizontal="center", vertical="center")
                  ws_write.column_dimensions["F"].width = 70
                  ws_write.row_dimensions[2].height = 60
      wb_write.save(file_put_write)

      subprocess.Popen([file_put_write], shell = True)



@eel.expose
def resultfileopen():
      """Открытие файла после обработки"""
      # Exposing the random_python function to javascript   
      subprocess.Popen([file_put], shell = True)

#поиск по реестру
@eel.expose
def start_search_js():
      ch_allcheck=eel.my_checkbox_function()()
      print (ch_allcheck)
      if ch_allcheck == "OK" :
            ch_citilink=eel.check_citilink()()
            ch_regard=eel.check_regard()()
            ch_dns=eel.check_dns()()
            if not os.path.exists("index_kp.txt"):
                  open("index_kp.txt", 'w').close()
            
            index_kp=read_file_index_kp() 
            if index_kp:
                  list_all_search=index_kp
            else:
                  list_all_search=["Запрос КП1","Запрос КП2","Запрос КП3","Запрос КП4","Запрос КП5","Запрос КП6","Запрос КП7","Запрос КП8"]
                  #list_all_search=["Запрос КП1"]

            for lname in list(list_all_search):                                    
                  if ch_citilink == "citilink":
                  #(путь к файлу, имя листа, имя  резервного листа, буква основной колонки, буква резервной колонки, имя столбца основного,имя столца резервного, имя функции поиска цены, буква столбца записи, имя столбца записи)
                        read2_xlsx(file_put,lname,"Реестр","h","e","Примечание","Наименование необходимых позиций",citilink,lname,"i","Ситилинк")                        
                  if ch_regard == "regard":
                        read2_xlsx(file_put,lname,"Реестр","h","e","Примечание","Наименование необходимых позиций",regard,lname,"j","Регард")
                  if ch_dns == "dns":
                        read2_xlsx(file_put,lname,"Реестр","h","e","Примечание","Наименование необходимых позиций",dnsshop,lname,"k","ДНС")                        
                  list_all_search.remove(lname)
                  write_file_index_kp(list_all_search)      

#поиск по одному
@eel.expose
def start_search_js_for_one(input_text):
      if input_text:
            global citilink_toex,regard_toex,dns_toex
            citilink_toex=[]
            regard_toex=[]
            dns_toex=[]
            #citilink_toex=""
            #regard_toex=""
            #dns_toex=""
            ch_allcheck=eel.my_checkbox_function()()
            print (ch_allcheck)
            if ch_allcheck == "OK" :
                  ch_citilink=eel.check_citilink()()
                  ch_regard=eel.check_regard()()
                  ch_dns=eel.check_dns()()
                  if ch_citilink == "citilink":
                        citilink_toex=search_of_one(input_text,citilink)
                        #citilink_toex.append(citilink_toexf)
                  if ch_regard == "regard":
                        regard_toex=search_of_one(input_text,regard)
                        #regard_toex.append(regard_toexf)
                  if ch_dns == "dns":
                        dns_toex=search_of_one(input_text,dnsshop)
                        #dns_toex.append(dns_toexf)
      else:
            eel.js_alert()
       


#обработка вывод на страницу результата
@eel.expose
def search_of_one(input_text, function_search):
      eel.js_wait()
      print(input_text, function_search)
      result_to_site=""
      result_to_excel=[]
      driver.start()
      buf = runSearchBySite(input_text, function_search(driver))
      result_to_site=f"✓ Результат поиска по магазину {function_search.__name__}"+"<br>"
      
      for i in buf:
            print(f"Вывод функции поиска {i.name} - Цена: {i.price}\n")
            result_to_site+="<a href=\""+i.url+"\" target=\"_blank\">"+"<font color=\"blue\">"+"• "+i.name+"</font>"+" - Цена: "+i.price+" ₽"+" </a>"+"<br>"
            result_to_excel.append([i.name,i.price])
            
      print(f"{result_to_site=}")
      driver.stop()
      eel.my_output_for_one(result_to_site)
      eel.js_gotovo()

      return result_to_excel






def website(mode):
      """Функция запуска поиска (параметр "reestr" - поиск по файлу реестре, параметр "one" - поиск по единичной позиции)"""
      os.environ["GOOGLE_API_KEY"] = "no"
      os.environ["GOOGLE_DEFAULT_CLIENT_ID"] = "no"
      os.environ["GOOGLE_DEFAULT_CLIENT_SECRET"] = "no"
      if mode=="reestr":
            eel.init("web")
      elif mode=="one":
            eel.init("web_for_one")
      
# Start the index.html file
      eel.browsers.set_path("chrome", "chrome-win/chrome.exe")
      eel.start("index.html",size=(1400, 1000))